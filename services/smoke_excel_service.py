from __future__ import annotations
from datetime import date, timedelta
from typing import Optional, List, Tuple

import os
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

BLOCK_COLS = 5                 # Pořadí, Druh, Poznámka, Dávka, Směna
ROWS_PER_SMOKER = 7
WEEKDAYS_7 = ["Pondělí","Úterý","Středa","Čtvrtek","Pátek","Sobota","Neděle"]

# ---------- util: bezpečné porovnání textu ----------
def _norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    s = " ".join(s.split()).lower()
    return s

def _is_header_label(v) -> bool:
    # přijme "Pořadové číslo", "poradove cislo", s mezerou na konci apod.
    return _norm(v).startswith("poradove cislo")

def _fmt_cz_date(d: date) -> str:
    return f"{d.day:02d}.{d.month:02d}.{d.year}"

# ---------- robustní výběr listu ----------
def _pick_worksheet(tpl_path: str, sheet_name: Optional[str]) -> Worksheet:
    wb = load_workbook(tpl_path, data_only=True, read_only=False)

    # 1) jméno listu, pokud je zadáno a existuje
    if sheet_name:
        try:
            ws = wb[sheet_name]
            if isinstance(ws, Worksheet):
                return ws
        except KeyError:
            pass

    # 2) první viditelný pracovní list
    for ws in getattr(wb, "worksheets", []):
        if getattr(ws, "sheet_state", "visible") == "visible":
            return ws

    # 3) jakýkoli pracovní list
    if getattr(wb, "worksheets", []):
        return wb.worksheets[0]

    # 4) nic nenalezeno
    names = list(getattr(wb, "sheetnames", []))
    raise ValueError(f"V šabloně se nepodařilo najít žádný pracovní list. Dostupné listy: {names or '— žádné —'}")

# ---------- autodetekce layoutu (počet udíren + řádky dní) ----------
def _detect_layout(ws: Worksheet) -> Tuple[int, List[int]]:
    """
    Vrátí (smokers, header_rows).
    smokers … počet bloků po 5 sloupcích (>=1)
    header_rows … řádky, kde začínají hlavičky (jeden řádek na den)
    """
    header_rows: List[int] = []
    smokers = 0
    max_row = ws.max_row or 200
    for r in range(1, max_row + 1):
        # počítej kolik bloků (1 + k*5) na řádku začíná "Pořadové číslo"
        count = 0
        for k in range(1, 20):  # bezpečný strop
            col = 1 + (k - 1) * BLOCK_COLS
            if _is_header_label(ws.cell(r, col).value):
                count += 1
            else:
                break
        if count > 0:
            header_rows.append(r)
            smokers = max(smokers, count)
    if smokers == 0 or not header_rows:
        raise ValueError("V šabloně jsem nenašel řádek s hlavičkami ('Pořadové číslo' v blocích).")
    return smokers, header_rows

# ---------- bezpečné zapsání (ignoruje merged read-only) ----------
def _safe_set(ws: Worksheet, row: int, col: int, value) -> None:
    try:
        ws.cell(row, col).value = value
    except AttributeError:
        # MergedCell je read-only (není to levý-horní roh spojení) → přeskoč
        pass

def _display_name(rw) -> str:
    rc_val = rw.get("rc")
    name_val = rw.get("polotovar_nazev")

    # Bezpečné "očištění" hodnot tak, aby se nezobrazovalo "nan", "<NA>", "None", apod.
    def _clean(v) -> str:
        if v is None:
            return ""
        try:
            # pokud je k dispozici pandas, korektně detekuj NaN/NA
            if pd.isna(v):
                return ""
        except Exception:
            pass
        s = str(v).strip()
        if s.lower() in {"", "nan", "<na>", "nat", "none", "null"}:
            return ""
        return s

    rc = _clean(rc_val)
    name = _clean(name_val)

    if rc and name:
        return f"400-{rc} - {name}"
    if rc:
        return f"400-{rc}"
    return name

    
def write_smoke_plan_excel(path: str,
                           plan_df: pd.DataFrame,
                           week_monday: Optional[date] = None,
                           sheet_name: Optional[str] = None,
                           template_path: Optional[str] = None) -> None:
    # ---- vstupní DF: striktně DataFrame + doplněné sloupce ----
    if not isinstance(plan_df, pd.DataFrame):
        raise TypeError("plan_df must be a pandas DataFrame")

    df = plan_df.copy()
    for c in ["datum","udirna","pozice","rc","polotovar_nazev","mnozstvi","jednotka","davka","shift","poznamka"]:
        if c not in df.columns:
            df[c] = pd.NA
    df["datum"] = pd.to_datetime(df["datum"], errors="coerce").dt.date

    if week_monday is None:
        valid_dates = [d for d in df["datum"] if d is not None]
        week_monday = (min(valid_dates) if valid_dates else date.today())

    # ---- načti šablonu + vyber list ----
    from services.smoke_paths import smoke_template_path
    tpl = template_path or str(smoke_template_path())
    if not os.path.exists(tpl):
        raise FileNotFoundError(f"Šablona nenalezena: {tpl}")
    ws = _pick_worksheet(tpl, sheet_name)

    # ---- auto-detekce: počet udíren a řádky dnů ----
    smokers_in_template, header_rows = _detect_layout(ws)
    # počet dnů – vezmeme, kolik bloků je v šabloně (obvykle 6 nebo 7)
    day_count = min(len(header_rows), len(WEEKDAYS_7))
    header_rows = header_rows[:day_count]

    # ---- vyčisti data + nastav nadpisy dnů ----
    for day_idx, hdr in enumerate(header_rows):
        title_row = max(1, hdr - 4)
        day_date = week_monday + timedelta(days=day_idx)
        _safe_set(ws, title_row, 1, f"{WEEKDAYS_7[day_idx]} {_fmt_cz_date(day_date)}")

        for s in range(smokers_in_template):
            start = 1 + s * BLOCK_COLS
            name_c, note_c, dose_c, = start + 1, start + 2, start + 3
            for i in range(1, ROWS_PER_SMOKER + 1):
                rr = hdr + i
                _safe_set(ws, rr, name_c, "")
                _safe_set(ws, rr, note_c, "")   # vždy prázdné
                _safe_set(ws, rr, dose_c, "")
                # SHIFT NEPÍŠEME (ve vzoru je často vertikálně sloučený) → zůstane prázdný

    # ---- zápis dat (ignorujeme udírny/pozice, které v šabloně nejsou) ----
    df["_udirna"] = pd.to_numeric(df["udirna"], errors="coerce").fillna(0).astype(int)
    df["_pozice"] = pd.to_numeric(df["pozice"], errors="coerce").fillna(0).astype(int)

    def _safe_str(val) -> str:
        try:
            if pd.isna(val):
                return ""
        except Exception:
            pass
        return "" if val is None else str(val)

        # --- nahraď tuto funkci (už bez fallbacku na množství) ---
    def _dose_from_row(rw) -> str:
        return _safe_str(rw.get("davka"))

    # --- přidej tuto novou pomocnou funkci ---
    def _note_from_row(rw) -> str:
        qty = rw.get("mnozstvi")
        unit = _safe_str(rw.get("jednotka"))
        if pd.notna(qty) and str(qty) != "":
            return f"{qty} {unit}".strip()
        # když není číslo, ale je jednotka, dej aspoň jednotku; jinak prázdné
        return unit if unit else ""


    # mapa datumů v rámci týdne (podle počtu dnů v šabloně)
    day_dates = {i: (week_monday + timedelta(days=i)) for i in range(day_count)}

    for day_idx, hdr in enumerate(header_rows):
        cur_date = day_dates.get(day_idx)
        day_rows = df[df["datum"] == cur_date] if cur_date else df.iloc[0:0]

        for s in range(1, smokers_in_template + 1):
            start = 1 + (s - 1) * BLOCK_COLS
            name_c, note_c, dose_c = start + 1, start + 2, start + 3

            for pos in range(1, ROWS_PER_SMOKER + 1):
                rec = day_rows[(day_rows["_udirna"] == s) & (day_rows["_pozice"] == pos)]
                if rec.empty:
                    continue
                rw = rec.iloc[0]

                display_name = _display_name(rw)
                note = _note_from_row(rw)     # <<< nově
                dose = _dose_from_row(rw)     # jen davka, bez fallbacku

                rr = hdr + pos
                _safe_set(ws, rr, name_c, display_name)
                _safe_set(ws, rr, note_c, note)   
                _safe_set(ws, rr, dose_c, dose)   # dávka jen když je, jinak prázdné


    # uložit kopii šablony s doplněnými daty
    ws.parent.save(path)
