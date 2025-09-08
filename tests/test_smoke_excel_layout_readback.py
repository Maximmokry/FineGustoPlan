# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook

from services.smoke_excel_service import write_smoke_plan_excel


def test_excel_layout_readback(tmp_path):
    week = pd.Timestamp("2025-09-08").date()  # pondělí
    # připrav miniplán na pondělí (ostatní dny writer vypíše prázdné bloky)
    plan_df = pd.DataFrame([
        {"datum": "2025-09-08", "den": "Pondělí", "udirna": 1, "pozice": 1,
         "polotovar_nazev": "Šunka", "mnozstvi": 200, "jednotka": "kg"},
        {"datum": "2025-09-08", "den": "Pondělí", "udirna": 2, "pozice": 1,
         "polotovar_nazev": "Krkovička", "mnozstvi": 150, "jednotka": "kg"},
    ])

    out = tmp_path / "plan_uzeni_layout.xlsx"
    write_smoke_plan_excel(str(out), plan_df, week)

    # Otevři první list (název listu není garantovaný)
    wb = load_workbook(out)
    ws = wb.worksheets[0]

    # 1) Najdi nadpis dne (writer zapisuje do sloupce A na řádek hdr-4, např. "Pondělí 08.09.2025")
    title_found = False
    for r in range(1, 200):
        v = str(ws.cell(r, 1).value or "")
        if v.startswith("Pondělí "):
            assert "08.09.2025" in v  # CZ formát data
            title_found = True
            break
    assert title_found, "Nadpis dne 'Pondělí DD.MM.YYYY' nebyl nalezen."

    # 2) Najdi řádek hlavičky bloků (kde je v A buňce text 'Pořadové číslo')
    hdr_row = None
    for r in range(1, 200):
        if (ws.cell(r, 1).value or "").strip() == "Pořadové číslo":
            hdr_row = r
            break
    assert hdr_row is not None, "Nenalezen řádek s hlavičkou 'Pořadové číslo'."

    # 3) Zkontroluj první datový řádek (pozice 1) – jména a součtové poznámky
    # Blok udíren je po 5 sloupcích: start=1 → Druh=2, Poznámka=3; start=6 → Druh=7, Poznámka=8
    row = hdr_row + 1
    # U1 / poz1
    assert (ws.cell(row, 2).value or "").strip() == "Šunka"
    assert (str(ws.cell(row, 3).value or "").strip()) == "200 kg"
    # U2 / poz1
    assert (ws.cell(row, 7).value or "").strip() == "Krkovička"
    assert (str(ws.cell(row, 8).value or "").strip()) == "150 kg"
