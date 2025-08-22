# tests/test_semi_excel_service_io.py
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import importlib

from services import semi_excel_service as ses

def test_ensure_output_semis_excel(tmp_path, monkeypatch):
    out_file = tmp_path / "polotovary.xlsx"
    # přesměruj OUTPUT_SEMI_EXCEL
    from services import paths as sp
    monkeypatch.setattr(sp, "OUTPUT_SEMI_EXCEL", str(out_file))

    df_main = pd.DataFrame([
        {"datum":"2025-03-01","polotovar_sk":"10","polotovar_rc":"A1","nazev":"Těsto","potreba":2,"jednotka":"kg","vyrobeno":False},
        {"datum":"2025-03-01","polotovar_sk":"20","polotovar_rc":"B2","nazev":"Omáčka","potreba":1.5,"jednotka":"l","vyrobeno":True},
    ])
    df_det = pd.DataFrame([
        {"datum":"2025-03-01","polotovar_sk":"10","polotovar_rc":"A1","final_rc":"F1","final_nazev":"Pizza","mnozstvi":2,"jednotka":"kg"},
    ])

    ses.ensure_output_semis_excel(df_main, df_det)

    assert out_file.exists()
    wb = load_workbook(out_file)
    assert set(wb.sheetnames) >= {"Prehled","Detaily","Polotovary"}
    # headery na Polotovary
    ws = wb["Polotovary"]
    exp = ["Datum","SK","Reg.č.","Polotovar","Množství","","Vyrobeno","Poznámka"]
    got = [(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert got == exp
