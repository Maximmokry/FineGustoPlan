# tests/test_uc10_export_polotovary.py
import os
from pathlib import Path
import pandas as pd
import pytest
from openpyxl import load_workbook

# headless režim (kdyby se něco dotklo Qt)
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

# projektové moduly
import services.paths as sp
from services.semi_excel_service import ensure_output_semis_excel


OUT = Path("test_uc10_polotovary.xlsx")


@pytest.fixture(autouse=True)
def _isolate_out(monkeypatch):
    # přesměrování cesty výstupu na testovací soubor
    monkeypatch.setattr(sp, "OUTPUT_SEMI_EXCEL", OUT, raising=False)
    if OUT.exists():
        OUT.unlink()
    yield
    if OUT.exists():
        OUT.unlink()


def _df_main():
    # dva řádky stejného polotovaru (jiná data), aby byly 2 mastery
    return pd.DataFrame(
        [
            {
                "datum": "2025-06-01",
                "polotovar_sk": "300",
                "polotovar_rc": "88",
                "polotovar_nazev": "Polotovar A",
                "potreba": 10,
                "jednotka": "kg",
                "vyrobeno": False,
            },
            {
                "datum": "2025-06-02",
                "polotovar_sk": "300",
                "polotovar_rc": "88",
                "polotovar_nazev": "Polotovar A",
                "potreba": 5,
                "jednotka": "kg",
                "vyrobeno": False,
            },
        ]
    )


def _df_details():
    # ke každému masteru dáme 2 podřádky
    return pd.DataFrame(
        [
            {
                "datum": "2025-06-01",
                "polotovar_sk": "300",
                "polotovar_rc": "88",
                "vyrobek_sk": "401",
                "vyrobek_rc": "1",
                "vyrobek_nazev": "Výrobek X",
                "mnozstvi": 6,
                "jednotka": "kg",
            },
            {
                "datum": "2025-06-01",
                "polotovar_sk": "300",
                "polotovar_rc": "88",
                "vyrobek_sk": "402",
                "vyrobek_rc": "2",
                "vyrobek_nazev": "Výrobek Y",
                "mnozstvi": 4,
                "jednotka": "kg",
            },
            {
                "datum": "2025-06-02",
                "polotovar_sk": "300",
                "polotovar_rc": "88",
                "vyrobek_sk": "401",
                "vyrobek_rc": "1",
                "vyrobek_nazev": "Výrobek X",
                "mnozstvi": 3,
                "jednotka": "kg",
            },
            {
                "datum": "2025-06-02",
                "polotovar_sk": "300",
                "polotovar_rc": "88",
                "vyrobek_sk": "402",
                "vyrobek_rc": "2",
                "vyrobek_nazev": "Výrobek Y",
                "mnozstvi": 2,
                "jednotka": "kg",
            },
        ]
    )


# ------------------------------------------------------------------
# UC10 – Polotovary list: hlavička + master/child + poznámka u mastera
# ------------------------------------------------------------------
def test_uc10_polotovary_sheet_structure_and_rows():
    ensure_output_semis_excel(_df_main(), _df_details())
    assert OUT.exists(), "Soubor s polotovary se měl vytvořit."

    wb = load_workbook(OUT)
    # 1) musí existovat listy
    assert "Prehled" in wb.sheetnames
    assert "Detaily" in wb.sheetnames
    assert "Polotovary" in wb.sheetnames

    ws = wb["Polotovary"]

    # 2) hlavička přesně dle specifikace (6. sloupec je skutečně prázdný text "")
    header = [ws.cell(1, c).value for c in range(1, 9)]
    assert header == ["Datum", "SK", "Reg.č.", "Polotovar", "Množství", None, "Vyrobeno", "Poznámka"]

    # 3) první master řádek – kontrola polí a umístění jednotky do sloupce 6
    #    (řádek 2 podle implementace)
    m1 = [ws.cell(2, c).value for c in range(1, 9)]
    assert str(m1[0]).startswith("2025-06-01"), "Datum masteru 1"
    assert m1[1] == "300" and str(m1[2]) == "88"
    assert m1[3] == "Polotovar A"
    assert float(m1[4]) == 10.0
    assert m1[5] == "kg", "Jednotka musí být ve 6. sloupci (bez hlavičky)"
    assert m1[6] in (False, 0), "Vyrobeno masteru má být False"
    # Poznámka se doplní až pokud existují děti – ověříme po vložení dětí níže

    # 4) podřádky pod prvním masterem (začínají šipkou ↳ ve 4. sloupci)
    #    očekáváme 2 děti:
    c1 = [ws.cell(3, c).value for c in range(1, 9)]
    c2 = [ws.cell(4, c).value for c in range(1, 9)]
    for child in (c1, c2):
        assert child[0] in (None, ""), "Datum u dítěte je prázdné"
        assert str(child[3]).startswith("↳ "), "Podřádek musí mít prefix '↳ ' v názvu"
        # E = množství, F = jednotka
        assert child[4] is not None
        assert child[5] in ("kg", ""), "Jednotka dítěte ve 6. sloupci"

    # 5) poznámka "(obsahuje rozpad)" u masteru s dětmi – ve sloupci 8
    assert ws.cell(2, 8).value == "(obsahuje rozpad)"

    # 6) druhý master a jeho děti (řádky pokračují)
    # master 2
    m2 = [ws.cell(5, c).value for c in range(1, 9)]
    assert str(m2[0]).startswith("2025-06-02")
    assert float(m2[4]) == 5.0
    assert m2[5] == "kg"
    # děti 2. masteru
    d1 = [ws.cell(6, c).value for c in range(1, 9)]
    d2 = [ws.cell(7, c).value for c in range(1, 9)]
    assert str(d1[3]).startswith("↳ ")
    assert str(d2[3]).startswith("↳ ")
    assert ws.cell(5, 8).value == "(obsahuje rozpad)"


# ------------------------------------------------------------------
# UC10 – Bez detailů: list existuje, Polotovary nemá podřádky, ale vše nespadne
# ------------------------------------------------------------------
def test_uc10_no_details_sheet_still_valid_output():
    # varianta A: df_details=None
    ensure_output_semis_excel(_df_main().iloc[:1], df_details=None)
    wb = load_workbook(OUT)
    assert "Prehled" in wb.sheetnames
    assert "Detaily" in wb.sheetnames, "I bez detailů se má zapsat prázdný list s hlavičkou"
    assert "Polotovary" in wb.sheetnames

    ws = wb["Polotovary"]
    header = [ws.cell(1, c).value for c in range(1, 9)]
    assert header[5] == None 

    # první data řádku (master) – nesmí následovat děti s '↳'
    m = [ws.cell(2, c).value for c in range(1, 9)]
    assert str(m[3]) == "Polotovar A"
    # další řádek (3) by neměl existovat jako '↳'
    next_val = ws.cell(3, 4).value
    assert not (isinstance(next_val, str) and next_val.startswith("↳ ")), "Bez detailů nesmí být podřádky"


# ------------------------------------------------------------------
# UC10 – Idempotence: opakovaný zápis neudělá rozbitou hlavičku ani duplicity
# ------------------------------------------------------------------
def test_uc10_idempotent_multiple_writes():
    ensure_output_semis_excel(_df_main(), _df_details())
    ensure_output_semis_excel(_df_main(), _df_details())  # znovu

    wb = load_workbook(OUT)
    ws = wb["Polotovary"]

    # hlavička drží
    header = [ws.cell(1, c).value for c in range(1, 9)]
    assert header == ["Datum", "SK", "Reg.č.", "Polotovar", "Množství", None, "Vyrobeno", "Poznámka"]

    # očekávaný počet řádků: 1 hlavička + 2 mastery + 4 děti = 7
    assert ws.max_row == 7, "Opakovaný zápis nesmí vytvářet další kopie řádků"
