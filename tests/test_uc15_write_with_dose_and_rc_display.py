# -*- coding: utf-8 -*-
"""
UC15 — Obsah Excelu: Druh, Poznámka (součet množství), Dávka
- Proč: obsluha potřebuje v Excelu srozumitelné hodnoty.
- Očekávání: 'Druh' = "400-<rc> - <název>" (když je rc), 'Poznámka' = "<množství> <MJ>", 'Dávka' = přesný text.
"""
from __future__ import annotations
from datetime import date
from pathlib import Path
import pandas as pd
import pytest
from openpyxl import load_workbook

from services.smoke_plan_service import SmokePlan
from services.smoke_excel_service import write_smoke_plan_excel
from tests._smoke_test_utils import create_smoke_template, open_xlsx

def _plan_df_with_one_item(week: date) -> pd.DataFrame:
    plan = SmokePlan(week)
    # vložíme jednu položku na Po, Udírna 1, Pozice 1
    from services.smoke_plan_service import Item, Slot
    plan.place(Item(
        polotovar_id_base="rc88|Hotový výrobek",
        polotovar_nazev="Hotový výrobek",
        mnozstvi=7.0,
        jednotka="kg",
        poznamka=None,
        meat_type=None,
    ), Slot(day_idx=0, smoker_idx=0, row_idx=0))
    df = plan.to_dataframe()
    # doplníme sloupce pro export
    df["davka"] = None
    df["shift"] = None
    df["poznamka"] = None
    # nastavíme 'rc' a dávku na přesné buňce (Po, U1, poz1)
    mask = (df["datum"].dt.date == week) & (df["udirna"] == 1) & (df["pozice"] == 1)
    df.loc[mask, "rc"] = "88"
    df.loc[mask, "davka"] = "60 min / 80 °C"
    return df

def _plan_df_with_one_item(week):
    from services.smoke_plan_service import SmokePlan, Item, Slot
    import pandas as pd

    plan = SmokePlan(week)
    plan.place(Item(
        polotovar_id_base="rc88|Hotový výrobek",
        polotovar_nazev="Hotový výrobek",
        mnozstvi=7.0,
        jednotka="kg",
        poznamka=None,
        meat_type=None,
    ), Slot(day_idx=0, smoker_idx=0, row_idx=0))

    df = plan.to_dataframe()
    # doplň sloupce pro export
    for c in ["rc", "davka", "shift", "poznamka"]:
        if c not in df.columns:
            df[c] = None

    # mask bez .dt – 'datum' je python date
    mask = (df["datum"] == week) & (df["udirna"] == 1) & (df["pozice"] == 1)
    df.loc[mask, "rc"] = "88"
    df.loc[mask, "davka"] = "60 min / 80 °C"
    return df


def test_uc15_display_fields_in_excel(tmp_path):
    print("UC15: Excel má správně 'Druh', 'Poznámka' a 'Dávku' pro první slot.")
    from tests._smoke_test_utils import create_smoke_template, open_xlsx
    from services.smoke_excel_service import write_smoke_plan_excel
    from datetime import date

    template = create_smoke_template(tmp_path / "tpl.xlsx")
    out = tmp_path / "plan.xlsx"
    week = date(2025, 9, 15)

    df = _plan_df_with_one_item(week)
    write_smoke_plan_excel(str(out), df, week_monday=week, sheet_name=None, template_path=str(template))

    wb = open_xlsx(out)
    ws = wb.active

    # hlavička v řádku 6 → data od 7; U1 blok: Druh sl.2, Poznámka sl.3, Dávka sl.4
    row = 7
    druh = ws.cell(row, 2).value
    poznamka = ws.cell(row, 3).value
    davka = ws.cell(row, 4).value

    assert druh == "400-88 - Hotový výrobek"
    assert str(poznamka).strip() == "7.0 kg"
    assert davka == "60 min / 80 °C"
