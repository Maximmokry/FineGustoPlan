# tests/test_uc2_semis.py
import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

import services.paths as sp
import services.semi_excel_service as ses
import gui.results_semis_window as rsw

# ---------------------------
# Headless / CI safe režim
# ---------------------------
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
os.environ.setdefault("FINEGUSTO_HEADLESS", "1")  # viz doporučení pro app (pokud je implementováno)

# Testovací výstup – nestrkat do ostrého polotovary.xlsx
TEST_FILE = Path("test_polotovary.xlsx")


@pytest.fixture(autouse=True)
def _clean_test_file(monkeypatch):
    # přesměruj cesty na testovací soubor (pro oba moduly, které je používají)
    monkeypatch.setattr(sp, "OUTPUT_SEMI_EXCEL", TEST_FILE)
    monkeypatch.setattr(rsw, "OUTPUT_SEMI_EXCEL", TEST_FILE)

    # minimalizace interakcí s GUI (popupy → no-op)
    monkeypatch.setattr(rsw.sg, "popup", lambda *a, **k: None)
    monkeypatch.setattr(rsw.sg, "popup_error", lambda *a, **k: None)

    if TEST_FILE.exists():
        TEST_FILE.unlink()
    yield
    if TEST_FILE.exists():
        TEST_FILE.unlink()


# ---------------------------
# Pomocné datové továrny
# ---------------------------
def make_df_main(mnozstvi=100, date_str="2025-05-12", sk="300", rc="10", nm="Uzený polotovar"):
    return pd.DataFrame(
        [
            {
                "datum": date_str,
                "polotovar_sk": sk,
                "polotovar_rc": rc,
                "polotovar_nazev": nm,
                "potreba": mnozstvi,
                "jednotka": "kg",
                "vyrobeno": False,
            }
        ]
    )


def make_df_det(mnozstvi=50, date_str="2025-05-12", sk="300", rc="10", nm="Hotový výrobek"):
    return pd.DataFrame(
        [
            {
                "datum": date_str,
                "polotovar_sk": sk,
                "polotovar_rc": rc,
                "vyrobek_sk": "400",
                "vyrobek_rc": "77",
                "vyrobek_nazev": nm,
                "mnozstvi": mnozstvi,
                "jednotka": "kg",
            }
        ]
    )


# ---------------------------
# UC2-1: Základní výpočet a zápis 3 listů + hlavička Polotovary
# ---------------------------
def test_writer_creates_sheets_and_header_cell_types():
    df_main = make_df_main(5)
    df_det = make_df_det(5)

    ses.ensure_output_semis_excel(df_main, df_det)

    assert TEST_FILE.exists(), "Soubor s polotovary se měl vytvořit."
    wb = load_workbook(TEST_FILE)
    assert set(["Prehled", "Detaily", "Polotovary"]).issubset(set(wb.sheetnames))

    ws = wb["Polotovary"]
    # hlavička
    header = [ws.cell(row=1, column=i).value for i in range(1, 9)]
    assert header[:5] == ["Datum", "SK", "Reg.č.", "Polotovar", "Množství"]
    assert header[5] in ("", None)              # ← tolerantní k openpyxl
    assert header[6:] == ["Vyrobeno", "Poznámka"]


# ---------------------------
# UC2-1b: Detaily – prázdný list má správné hlavičky
# ---------------------------
def test_writer_creates_empty_details_when_none():
    df_main = make_df_main(5)
    ses.ensure_output_semis_excel(df_main, None)
    wb = load_workbook(TEST_FILE)
    assert "Detaily" in wb.sheetnames
    ws = wb["Detaily"]
    headers = [ws.cell(1, i).value for i in range(1, 9)]
    assert headers == [
        "datum",
        "polotovar_sk",
        "polotovar_rc",
        "vyrobek_sk",
        "vyrobek_rc",
        "vyrobek_nazev",
        "mnozstvi",
        "jednotka",
    ]


# ---------------------------
# UC2-2: Update množství – přepsání na aktuální stav (100 -> 200)
# ---------------------------
def test_update_quantity_overwrites_to_current_state():
    df_main = make_df_main(100)
    ses.ensure_output_semis_excel(df_main, None)

    pre = pd.read_excel(TEST_FILE, sheet_name="Prehled")
    assert int(float(pre.loc[0, "potreba"])) == 100
    assert not bool(pre.loc[0, "vyrobeno"])

    df_main2 = make_df_main(200)
    ses.ensure_output_semis_excel(df_main2, None)

    post = pd.read_excel(TEST_FILE, sheet_name="Prehled")
    assert int(float(post.loc[0, "potreba"])) == 200  # změna se propsala


# ---------------------------
# UC2-3: Weekly agregace – Po–Ne label a součet
# ---------------------------
def test_week_range_label_and_aggregate_weekly_sum():
    df = pd.DataFrame(
        [
            {
                "datum": "2025-05-12",
                "polotovar_sk": "300",
                "polotovar_rc": "10",
                "polotovar_nazev": "Uzený",
                "jednotka": "kg",
                "potreba": 5,
                "vyrobeno": False,
            },
            {
                "datum": "2025-05-15",
                "polotovar_sk": "300",
                "polotovar_rc": "10",
                "polotovar_nazev": "Uzený",
                "jednotka": "kg",
                "potreba": 3,
                "vyrobeno": False,
            },
        ]
    )
    ses.ensure_output_semis_excel(df, None)
    prehled = pd.read_excel(TEST_FILE, sheet_name="Prehled")
    col_k = rsw.find_col(prehled, ["vyrobeno"]) or "vyrobeno"

    g = rsw._aggregate_weekly(prehled, col_k)
    assert not g.empty
    # jeden součet 5+3
    assert float(g["potreba"].iloc[0]) == 8.0
    # label týdne Po–Ne (pomlčka "–" a aspoň 4 tečky v datových částech)
    label = str(g["datum"].iloc[0])
    assert "–" in label and label.count(".") >= 4


# ---------------------------
# UC2-4: Detaily – builder přidá podřádky
# ---------------------------
def test_build_rows_with_details_adds_child_rows():
    df_main = make_df_main(5, date_str="2025-05-12", sk="300", rc="10", nm="Uzený")
    # detail map klíč: (polotovar_sk, polotovar_rc, datum)
    detail_map = {
        ("300", "10", pd.to_datetime("2025-05-12").date()): [
            {"vyrobek_sk": "400", "vyrobek_rc": "77", "vyrobek_nazev": "Hotový", "mnozstvi": 5, "jednotka": "kg"}
        ]
    }
    rows, buy_map, rowkey_map = rsw._build_rows(
        df_main, detail_map, "vyrobeno", show_details=True, weekly_sum=False
    )
    # hlavička + hlavní řádek + detail řádek
    assert rows is not None and len(rows) >= 3
    # existuje tlačítko -SEMI- (neagreg.)
    assert any(k.startswith("-SEMI-") for k in buy_map.keys())


# ---------------------------
# UC2-5: GUI (headless) – klik v neagregovaném režimu označí vyrobeno=True a uloží
# ---------------------------
class _FakeWindow:
    def __init__(self, events):
        self._events = list(events)

    def read(self, timeout=None):
        if self._events:
            return self._events.pop(0)
        return (None, {})

    def close(self):
        pass

    def current_location(self):
        return (100, 100)


def test_open_semis_results_click_marks_true_and_saves(monkeypatch):
    # připrav Prehled
    df_main = make_df_main(10, date_str="2025-06-01", sk="300", rc="99", nm="Šunka")
    ses.ensure_output_semis_excel(df_main, None)

    # fake _create_window: použije builder řádků a vrátí FakeWindow s klikem na první '-SEMI-'
    def fake_create(df_main_in, detail_map_in, col_k, show_details, weekly_sum, location=None):
        rows, buy_map, rowkey_map = rsw._build_rows(df_main_in, detail_map_in, col_k, show_details, weekly_sum)
        first = next((k for k in buy_map if k.startswith("-SEMI-")), None)
        events = []
        if first:
            events.append((first, {}))
        events.append((None, {}))  # zavření
        return _FakeWindow(events), buy_map, rowkey_map

    monkeypatch.setattr(rsw, "_create_window", fake_create)

    # rekreace okna – jen přestav, žádné skutečné GUI
    monkeypatch.setattr(rsw, "recreate_window_preserving", lambda old_win, builder, col_key='-COL-', **kw: builder((100, 100)))

    rsw.open_semis_results()

    # ověř uložení vyrobeno=True
    df_after = pd.read_excel(TEST_FILE, sheet_name="Prehled")
    col_k = rsw.find_col(df_after, ["vyrobeno"]) or "vyrobeno"
    assert col_k in df_after.columns
    assert bool(df_after[col_k].iloc[0]) is True


# ---------------------------
# UC2-6: GUI (headless) – weekly režim, klik na agregovaný řádek označí všechny zdroje
# ---------------------------
def test_open_semis_results_weekly_click_marks_all_true(monkeypatch):
    import pandas as pd
    from gui.results_semis_window import _build_rows as _rows_builder  # jen pro získání klíčů
    import gui.results_semis_window as rsw
    import services.semi_excel_service as ses
    from tests.test_uc2_semis import _FakeWindow, TEST_FILE  # používá se ve zbytku souboru

    # 1) Připrav dva řádky ve stejném týdnu, se stejným SK/RC/Názvem/MJ
    df_main = pd.DataFrame(
        [
            {
                "datum": "2025-05-12",  # Po
                "polotovar_sk": "300",
                "polotovar_rc": "10",
                "polotovar_nazev": "Uzený",
                "jednotka": "kg",
                "potreba": 5,
                "vyrobeno": False,
            },
            {
                "datum": "2025-05-15",  # Čt (stejný týden)
                "polotovar_sk": "300",
                "polotovar_rc": "10",
                "polotovar_nazev": "Uzený",
                "jednotka": "kg",
                "potreba": 3,
                "vyrobeno": False,
            },
        ]
    )
    ses.ensure_output_semis_excel(df_main, None)

    # 2) Fake _create_window: 
    #   - poprvé (weekly_sum=False) vrátí klik na toggle -WEEKLY-
    #   - podruhé (weekly_sum=True) vrátí klik na první -WSEMI- a pak None (zavření)
    def fake_create(df_main_in, detail_map_in, col_k, show_details, weekly_sum, location=None):
        rows, buy_map, rowkey_map = rsw._build_rows(df_main_in, detail_map_in, col_k, show_details, weekly_sum)

        events = []
        if not weekly_sum:
            # jen přepnout do weekly; neklikej znovu na -WEEKLY- v weekly režimu!
            events.append(("-WEEKLY-", {"-WEEKLY-": True}))
        else:
            # v weekly režimu už skutečně klikneme na agregovaný řádek
            first = next((k for k in buy_map if k.startswith("-WSEMI-")), None)
            if first:
                events.append((first, {}))
            events.append((None, {}))  # zavřít

        return _FakeWindow(events), buy_map, rowkey_map

    # Rekreace okna – jen znovu postav, žádné skutečné GUI
    def fake_recreate(old_win, builder, col_key='-COL-', **kw):
        return builder((100, 100))

    monkeypatch.setattr(rsw, "_create_window", fake_create)
    monkeypatch.setattr(rsw, "recreate_window_preserving", fake_recreate)

    # Vypnout pop-upy
    monkeypatch.setattr(rsw.sg, "popup", lambda *a, **k: None)
    monkeypatch.setattr(rsw.sg, "popup_error", lambda *a, **k: None)

    # 3) Spusť okno (headless)
    rsw.open_semis_results()

    # 4) Ověření: weekly klik označil oba zdrojové řádky jako True
    out = pd.read_excel(TEST_FILE, sheet_name="Prehled")
    col_k = rsw.find_col(out, ["vyrobeno"]) or "vyrobeno"
    assert out[col_k].astype(bool).all(), "Všechny zdrojové řádky v weekly agregaci mají být True"
