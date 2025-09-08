# -*- coding: utf-8 -*-
from __future__ import annotations
from pathlib import Path
from typing import Optional
from datetime import date
from openpyxl import Workbook, load_workbook

BLOCK_COLS = 5  # Pořadí, Druh, Poznámka, Dávka, Směna

def create_smoke_template(path: Path, smokers: int = 4, days: int = 6, rows_per_smoker: int = 7) -> Path:
    """
    Vytvoří minimalistickou šablonu, kterou umí detekovat services.smoke_excel_service:
    - na řádcích hlaviček je ve sloupci 1 + k*5 text 'Pořadové číslo'
    - pod každou hlavičkou je alespoň rows_per_smoker řádků pro zápis
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"

    # řádky s hlavičkami (1 hlavička na den)
    start_row = 6
    step = rows_per_smoker + 10
    for d in range(days):
        hdr_row = start_row + d * step
        for k in range(smokers):
            col = 1 + k * BLOCK_COLS
            ws.cell(hdr_row, col).value = "Pořadové číslo"
        # necháme 4 řádky nad hlavičkou pro titulek (Pondělí DD.MM.YYYY atd.)
        # data se budou zapisovat na řádky hdr_row+1 .. hdr_row+rows_per_smoker

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path

def patch_smoke_paths(monkeypatch, tmpdir: Path, template: Path):
    """Přesměruj plánovací cesty do tmp a použij zadanou šablonu."""
    from services import smoke_paths as SP

    def _plan_uzeni_dir():
        out = tmpdir / "plan uzeni"
        out.mkdir(parents=True, exist_ok=True)
        return out

    def _smoke_template_path():
        return template

    def _smoke_plan_excel_path(monday: date):
        fname = f"plan_uzeni_{monday:%Y_%m_%d}.xlsx"
        return _plan_uzeni_dir() / fname

    monkeypatch.setattr(SP, "plan_uzeni_dir", _plan_uzeni_dir, raising=False)
    monkeypatch.setattr(SP, "smoke_template_path", _smoke_template_path, raising=False)
    monkeypatch.setattr(SP, "smoke_plan_excel_path", _smoke_plan_excel_path, raising=False)

def open_xlsx(path: Path):
    return load_workbook(str(path))
